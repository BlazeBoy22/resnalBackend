import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, Color, PatternFill


# Load the JSON data
with open('result14.json', 'r') as file:
    data = json.load(file)

# Create a new workbook and select the active worksheet
workbook = openpyxl.Workbook()
sheet = workbook.active

border_style = Border(left=Side(style='thin', color='000000'),
                      right=Side(style='thin', color='000000'),
                      top=Side(style='thin', color='000000'),
                      bottom=Side(style='thin', color='000000'))

subject_codes = [subject["subjectCode"] for subject in data[0]["results"]]
subject_names = [subject["subjectName"] for subject in data[0]["results"]]
def create_headers():
    # Prepare the headers for the Excel sheet
    headers = ["Name", "USN", "Section"]
    
    # print(subject_codes)
    for i in range(len(subject_codes)):
        headers.append(f'{subject_codes[i]} - {subject_names[i]}')

    col_num = 1
    for header_title in headers:
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header_title
        cell.font = Font(bold=True)
        cell.border = border_style
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if(col_num < 4):
            cell = sheet.cell(row=2, column=col_num)
            cell.border = border_style
            col_num += 1
        else:
            cell = sheet.cell(row=2, column=col_num)
            cell.value = "IA"
            cell.font = Font(bold=True)
            cell.border = border_style
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell = sheet.cell(row=2, column=col_num+1)
            cell.value = "EA"
            cell.font = Font(bold=True)
            cell.border = border_style
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell = sheet.cell(row=2, column=col_num+2)
            cell.value = "TOTAL"
            cell.font = Font(bold=True)
            cell.border = border_style
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell = sheet.cell(row=2, column=col_num+3)
            cell.value = "CLASS"
            cell.font = Font(bold=True)
            cell.border = border_style
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # for i in range(4):
            #     col_letter = get_column_letter(col_num + i)
            #     sheet.column_dimensions[col_letter].width = 15

            col_num += 4

    # Merge cells for each subject starting from D1
    for col_num in range(4, 3 + 4*(len(headers)-3) + 1, 4):
        start_col = col_num
        end_col = start_col + 3
        sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 20

create_headers()

# Load the JSON data
with open('result14.json', 'r') as file:
    data = json.load(file)

rows = []
for student in data:
    row = [student["name"], student["USN"], student["Section"]]
    for subject_code in subject_codes:
        flag = 0
        for subject in student["results"]:
            if subject_code == subject["subjectCode"]:
                flag = 1
                if(subject["total"] == None or subject["ea"] == None):
                    class1 = ""
                elif(int(subject["total"]) < 40 or int(subject["ea"]) < 18):
                    class1 = "F"
                elif(int(subject["total"]) < 50):
                    class1 = "P"
                elif(int(subject["total"]) < 60):
                    class1 = "SC"
                elif(int(subject["total"]) < 70):
                    class1 = "FC"
                else:
                    class1 = "FCD"
                row.extend([subject["ia"], subject["ea"], subject["total"], class1])

        if flag == 0:
            row.extend(["", "", "", ""])
            for subject in student["results"]:
                if subject["subjectCode"] not in subject_codes:
                    subject_codes.append(f'{subject["subjectCode"]}')
                    subject_names.append(f'{subject["subjectName"]}')
                    create_headers()

    rows.append(row)

# print(rows)

# Iterate over the data and write to the worksheet
for row_index, row_data in enumerate(rows, start=3):
    for col_index, cell_value in enumerate(row_data, start=1):
        sheet.cell(row=row_index, column=col_index, value=cell_value)
        cell = sheet.cell(row=row_index, column=col_index)
        if cell_value == "FCD":
            cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        elif cell_value == "FC":
            cell.fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')
        elif cell_value == "SC":
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        elif cell_value == "P":
            cell.fill = PatternFill(start_color='7600BC', end_color='7600BC', fill_type='solid')
        elif cell_value == "F":
            cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

sections = []
max_row = sheet.max_row
for col_num in range(7, (len(subject_codes)+1)*4, 4):
    # col_letter = get_column_letter(col_num)

    fcd_count = 0
    fc_count = 0
    sc_count = 0
    p_count = 0
    f_count = 0
    
    for i in range(2, max_row):
        cell = sheet.cell(row=i+1, column=col_num)
        if(cell.value == "FCD"):
            fcd_count += 1
        elif(cell.value == "FC"):
            fc_count += 1
        elif(cell.value == "SC"):
            sc_count += 1
        elif(cell.value == "P"):
            p_count += 1
        elif(cell.value == "F"):
            f_count += 1
        
        cell = sheet.cell(row=i+1, column=3)
        if cell.value not in sections:
            sections.append(cell.value)
    
    sheet.cell(row=max_row + 2, column=3, value="Total")
    sheet.cell(row=max_row + 2, column=col_num, value=fcd_count)
    sheet.cell(row=max_row + 2, column=col_num-1, value="FCD")
    sheet.cell(row=max_row + 3, column=col_num, value=fc_count)
    sheet.cell(row=max_row + 3, column=col_num-1, value="FC")
    sheet.cell(row=max_row + 4, column=col_num, value=sc_count)
    sheet.cell(row=max_row + 4, column=col_num-1, value="SC")
    sheet.cell(row=max_row + 5, column=col_num, value=p_count)
    sheet.cell(row=max_row + 5, column=col_num-1, value="P")
    sheet.cell(row=max_row + 6, column=col_num, value=f_count)
    sheet.cell(row=max_row + 6, column=col_num-1, value="F")

    for i in range(1,len(sections)+1):
        fcd_count = 0
        fc_count = 0
        sc_count = 0
        p_count = 0
        f_count = 0

        # print(i)
        # print(sections[i-1])
        # cell = sheet.cell(row=3, column=3)
        # print(cell.value)
        for j in range(2, max_row):
            cell = sheet.cell(row=j+1, column=3)
            if(cell.value == sections[i-1]):
                cell = sheet.cell(row=j+1, column=col_num)
                if(cell.value == "FCD"):
                    fcd_count += 1
                elif(cell.value == "FC"):
                    fc_count += 1
                elif(cell.value == "SC"):
                    sc_count += 1
                elif(cell.value == "P"):
                    p_count += 1
                elif(cell.value == "F"):
                    f_count += 1
        
        sheet.cell(row=max_row+2 + 6*i, column=3, value=sections[i-1])
        sheet.cell(row=max_row+2 + 6*i, column=col_num, value=fcd_count)
        sheet.cell(row=max_row+2 + 6*i, column=col_num-1, value="FCD")
        sheet.cell(row=max_row+2 + 6*i+1, column=col_num, value=fc_count)
        sheet.cell(row=max_row+2 + 6*i+1, column=col_num-1, value="FC")
        sheet.cell(row=max_row+2 + 6*i+2, column=col_num, value=sc_count)
        sheet.cell(row=max_row+2 + 6*i+2, column=col_num-1, value="SC")
        sheet.cell(row=max_row+2 + 6*i+3, column=col_num, value=p_count)
        sheet.cell(row=max_row+2 + 6*i+3, column=col_num-1, value="P")
        sheet.cell(row=max_row+2 + 6*i+4, column=col_num, value=f_count)
        sheet.cell(row=max_row+2 + 6*i+4, column=col_num-1, value="F")

# print(sections)

# Save the workbook
workbook.save("student_marks.xlsx")
