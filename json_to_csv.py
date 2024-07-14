import json
import pandas as pd

# Load the JSON data
with open('result14.json', 'r') as file:
    data = json.load(file)

# Prepare the headers for the Excel sheet
headers = ["Student Name", "Student USN", "Section"]
# subject_codes = [subject["subjectCode"] for subject in data[0]["results"]]

subject_codes = [
    "21CIV57",
    "21CS51",
    "21CS52",
    "21CS53",
    "21CS54",
    "21CSL55",
    "21CSL581",
    "21CSL582",
    "21RMI56"
]

for code in subject_codes:
    headers.extend([f"{code} ia", f"{code} ea", f"{code} total"])

# Create a DataFrame to hold the data
rows = []
for student in data:
    row = [student["name"], student["USN"], student["Section"]]
    # for subject in student["results"]:
    #     row.extend([subject["ia"], subject["ea"], subject["total"]])
    # rows.append(row)

    for subject_code in subject_codes:
        for subject in student["results"]:
            if subject_code == subject["subjectCode"]:
                if subject_code == "21CSL581":
                    row.extend([subject["ia"], subject["ea"], subject["total"]])
                    row.extend(["", "", ""])
                elif subject_code == "21CSL582":
                    row.extend(["", "", ""])
                    row.extend([subject["ia"], subject["ea"], subject["total"]])
                else:
                    row.extend([subject["ia"], subject["ea"], subject["total"]])

    rows.append(row)
            

# Create the DataFrame
df = pd.DataFrame(rows, columns=headers)

# Create an Excel writer object
with pd.ExcelWriter('output.xlsx', engine='openpyxl') as writer:
    # Write the DataFrame to Excel starting from row 3
    df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=1)

    # Access the workbook and sheet
    workbook = writer.book
    sheet = writer.sheets['Sheet1']

    # Add the headers and subheaders
    header_row = 1
    subheader_row = 2
    for col_num, header in enumerate(headers):
        sheet.cell(row=header_row, column=col_num + 1, value=header)
        sheet.cell(row=subheader_row, column=col_num + 1, value="")
        if header not in ["Student Name", "Student USN", "Section"]:
            if "ia" in header:
                sheet.cell(row=header_row, column=col_num + 1, value=header[:-3])
                sheet.cell(row=subheader_row, column=col_num + 1, value="ia")
            elif "ea" in header:
                sheet.cell(row=header_row, column=col_num + 1, value=header[:-3])
                sheet.cell(row=subheader_row, column=col_num + 1, value="ea")
            elif "total" in header:
                sheet.cell(row=header_row, column=col_num + 1, value=header[:-6])
                sheet.cell(row=subheader_row, column=col_num + 1, value="total")

    # Adjust the formatting as needed (e.g., bold headers, column widths, etc.)
    for cell in sheet[header_row]:
        cell.font = cell.font.copy(bold=True)
    for cell in sheet[subheader_row]:
        cell.font = cell.font.copy(bold=True)
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 15

print("Excel file has been created successfully.")

#-------------------------------------------------------------------------------

import openpyxl
from openpyxl.styles import Alignment

def merge_cells_with_value(sheet, merge_range):
    # Merge the cells
    sheet.merge_cells(merge_range)

# Load the workbook
file_path = 'output.xlsx'
wb = openpyxl.load_workbook(file_path)

# Select the worksheet (assuming it's the first sheet, you can change as needed)
sheet = wb.active

# Define multiple merge ranges
merge_ranges = ['D1:F1', 'G1:I1', 'J1:L1', 'M1:O1', 'P1:R1', 'S1:U1', 'V1:X1', 'Y1:AA1', 'AB1:AD1']

# Merge cells and set value '1' for each range
for merge_range in merge_ranges:
    merge_cells_with_value(sheet, merge_range)

start_col, end_col = 1, 30

for col in range(start_col, end_col + 1):
    cell = sheet.cell(row=1, column=col)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Save the workbook
wb.save(file_path)

print(f"Merged cells")

